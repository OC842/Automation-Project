import geopandas as gpd
import matplotlib.pyplot as plt
import pandas as pd

from lib.base.map_strings import MapStrings

import georaster

import matplotlib.image as mpimg
from matplotlib.markers import MarkerStyle

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from math import sqrt, tan, pi


#Add POI data to workbook using 'anchor_cell' as a base
def populate_spreadsheet(wb, poi_data, anchor_cell):


	ws = wb['POI Overview']

	titles = ['POI_ID', 'X', 'Y', 'Point Type', 'Confidence', 'Logger Profile', 'Distance to Address (m)']

	#Re-assigning address and material so they can be removed from 'poi_data'
	address = poi_data[7]
	material = poi_data[6]
	poi_data.pop(7)
	poi_data.pop(6)

	
	#Adding titles to worksheet
	for i in range(len(titles)):
		ws.cell(column=anchor_cell[0]+i, row=anchor_cell[1], value=titles[i])
	
	#Adding POI data to worksheet
	for i in range(len(titles)):
		ws.cell(column=anchor_cell[0]+i, row=anchor_cell[1]+1, value=poi_data[i])
		
		#Cell manipulation for better appearance
		if i == 3 or i == 5:
		
			ws.cell(column=anchor_cell[0]+i, row=anchor_cell[1]+1).number_format = '@'
			ws.cell(column=anchor_cell[0]+i, row=anchor_cell[1]+1).alignment = Alignment(horizontal="center", vertical="center")
			
		else:
			ws.cell(column=anchor_cell[0]+i, row=anchor_cell[1]+1).number_format = '0'
	
	#Cell manipulation for better appearance
	#ws.merge_cells('B3:G3')
	ws.merge_cells(start_row = anchor_cell[1]+2, start_column = anchor_cell[0]+1, end_row = anchor_cell[1]+2, end_column = anchor_cell[0]+6)
	#ws.merge_cells('B4:G4')
	ws.merge_cells(start_row = anchor_cell[1]+3, start_column = anchor_cell[0]+1, end_row = anchor_cell[1]+3, end_column = anchor_cell[0]+6)
	#ws.merge_cells('B5:D5')
	#ws.merge_cells(start_row = anchor_cell[1]+4, start_column = anchor_cell[0]+1, end_row = anchor_cell[1]+4, end_column = anchor_cell[0]+3)
	#ws.merge_cells('E4:G4')
	#ws.merge_cells(start_row = anchor_cell[1]+3, start_column = anchor_cell[0]+4, end_row = anchor_cell[1]+3, end_column = anchor_cell[0]+6)
	#ws.merge_cells('E5:G5')
	#ws.merge_cells(start_row = anchor_cell[1]+4, start_column = anchor_cell[0]+4, end_row = anchor_cell[1]+4, end_column = anchor_cell[0]+6)
	
	#Adding material and address headings
	ws.cell(column = anchor_cell[0], row = anchor_cell[1]+2).value = 'Material'
	ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+2).value = 'Address'
	ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+2).alignment = Alignment(horizontal="center", vertical="center")
	
	#Resizing column widths for better appearance (only do this the for first two POIs)
	if anchor_cell[1] == 1:
		column_width = [8, 10, 10, 10, 12, 14, 23]
		i = 0
		for i in range(len(column_width)):

			ws.column_dimensions[get_column_letter(anchor_cell[0] + i)].width = column_width[i]
				
			i += 1


	#Adding material and address values
	ws.cell(column = anchor_cell[0], row = anchor_cell[1]+3).value = material
	ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).value = address
	ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).alignment = Alignment(horizontal="center", vertical="center")
	
	
	#Change address font if address is too long for the box
	if len(address) > 120:
	
		ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).font = Font(size = 6)
		
	elif len(address) > 105:
	
		ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).font = Font(size = 7)
		
	elif len(address) > 90:
	
		ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).font = Font(size = 8)
		
	elif len(address) > 80:
	
		ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).font = Font(size = 9)
		
	elif len(address) > 70:
	
		ws.cell(column = anchor_cell[0]+1, row = anchor_cell[1]+3).font = Font(size = 10)
	
	


	#Finding map file
	plot_folder = r'lib\png'
	img_file = '{0}\\{1}.png'.format(plot_folder, str(int(poi_data[0])))
	
	#Adding map to worksheet
	img = openpyxl.drawing.image.Image(img_file)
	image_column = get_column_letter(anchor_cell[0])
	image_cell = str(image_column) + str(anchor_cell[1]+5)
	ws.add_image(img, image_cell)
	
	
	#Changing anchor cell ready for the next POI
	#If anchor cell has already been moved to I - Move it down 25 rows and back to column A
	if anchor_cell[0] > 1:
		anchor_cell = [1, anchor_cell[1] + 25]
	
	#Move cell along 8 columns from A to I
	else:
		anchor_cell = [anchor_cell[0] + 8, anchor_cell[1]]
		
		
	return anchor_cell


#Uses the DataCheck class to ensure all input data is within DMA
def data_within_outline(input_data, checker):

	output_data = []

	for i in range(len(input_data)):
	
		point = [input_data[i][0], input_data[i][1]]

		if checker.check_point(point, 10):
		
			output_data.append(input_data[i])
		
	return output_data


#Get address info from .shp file
def address_info(dma_outline):

	#Read file
	address_file = r'.shp\Addresses\Addresses.shp'
	address_data = gpd.read_file(address_file, dma_outline)
	
	#Extract info from dataframe into lists
	x_coords = address_data['X_COORDI_1'].tolist()
	y_coords = address_data['Y_COORDI_1'].tolist()
	addresses = address_data['LOCATION_D'].tolist()
	
	#Any address with these key words is not useful so shall be ignored
	non_addresses = ['ELECTRICITY SUB STATION', 'STREET RECORD', 'POST BOX', 'CCTV CAMERA IN OPENSPACE', 'TELECOMS APPARATUS', 'PUBLIC TELEPHONE']
	
	#Loop to filter address data
	address_list = []
	for i in range(len(address_data)):
	
		add_address = True
	
		#If address contains non address keyword then do not add it to 'address_list'
		for non_add in non_addresses:
			if non_add in str(addresses[i]):
				add_address = False
		
		#Extra check to remove entries such as 'SHELTER 8M FROM HIGH STREET'
		#A real address may contain 'SHELTER' so it cannot be added to 'non_addresses'
		#Wildcards could be incorporated into the non addresses check however that would add additional 
		# complexity which is not needed at this time
		first_address = str(addresses[i]).split(',')
		if 'SHELTER' in first_address[0] and 'M FROM' in first_address[0]:
			add_address = False
			
		if 'POND' in first_address[0] and 'M FROM' in first_address[0]:
			add_address = False
		
		#If address does not contain any keywords then add it to 'address_list'
		if add_address:
		
			address_list.append([x_coords[i], y_coords[i], addresses[i]])
			
	return address_list

#Get valve info from .shp file
def valve_info(dma_outline):

	#Read file
	valve_file = r'.shp\Valves\Valves.shp'
	valve_data = gpd.read_file(valve_file, dma_outline)
	
	#Extract info from dataframe into lists
	#Information about the subtype and current_op codes can be found in the 'codes' folder
	#'orientatio' is the orientation of the valves which is needed for mapping purposes
	valve_subtype = valve_data['subtype'].tolist()
	valve_orientation = valve_data['orientatio'].tolist()
	valve_status = valve_data['current_op'].tolist()
	
	valve_x = valve_data['POINT_X'].tolist()
	valve_y = valve_data['POINT_Y'].tolist()
	
	#Create valve info array
	valve_info = []
	for i in range(len(valve_subtype)):
	
		#1, 14 and 16 correspond to valves which have been abandoned, proposed or removed so are not currently active
		#More codes at 'codes\Valves'
		if not valve_subtype[i] in [1, 14, 16]:
		
			valve_info.append([valve_x[i], valve_y[i], valve_subtype[i], valve_status[i], valve_orientation[i]])
	
	
	return valve_info

#Get hydrant info from .shp file
def hydrant_info(dma_outline):

	#Read file
	hydrant_file = r'.shp\Hydrants\Hydrants.shp'
	hydrant_data = gpd.read_file(hydrant_file, dma_outline)
	
	#Extract info from dataframe into lists
	#Information about the subtype codes can be found in the 'codes' folder
	hydrant_subtype = hydrant_data['subtype'].tolist()

	hydrant_x = hydrant_data['POINT_X'].tolist()
	hydrant_y = hydrant_data['POINT_Y'].tolist()
	
	#Creating hydrant info array
	hydrant_info = []
	for i in range(len(hydrant_subtype)):
	
		#1, 5 and 6 correspond to valves which have been abandoned, proposed or removed so are not currently active
		#More codes at 'codes\Hydrants\Hydrant Sub-Type.txt'
		if not hydrant_subtype[i] in [1, 5, 6]:
		
			hydrant_info.append([hydrant_x[i], hydrant_y[i], hydrant_subtype[i]])
			
	return hydrant_info

#Get PRV info from .shp file
def prv_info(dma_outline):

	#Read file
	prv_file = r'.shp\PRVs\PRVs.shp'
	prv_data = gpd.read_file(prv_file, dma_outline)

	#Extract info from dataframe into lists
	#Information about the subtype codes can be found in the 'codes' folder
	#'orientatio' is the orientation of the valve which is needed for mapping purposes
	prv_subtype = prv_data['subtype'].tolist()
	prv_orientation = prv_data['orientatio'].tolist()

	prv_x = prv_data['POINT_X'].tolist()
	prv_y = prv_data['POINT_Y'].tolist()
	
	#Creating PRV info array
	prv_info = []
	for i in range(len(prv_subtype)):
	
		#1, 6 and 7 correspond to valves which have been abandoned, proposed or removed so are not currently active
		#More codes at 'codes\PRVs\PRV Sub-Type.txt'
		if not prv_subtype[i] in [1, 6, 7]:
		
			prv_info.append([prv_x[i], prv_y[i], prv_subtype[i], prv_orientation[i]])
			
	return prv_info
	

#Plot valve info
def plot_valve(ax, valve_info):

	#Plotting valves according to their GIS symbology
	#More information can be found at 'codes\Valves\Valve Colours.txt'
	if valve_info[2] in [2, 9, 10, 11, 12, 13, 18, 19]:
	
		#Default
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 4)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 3, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
		
		
	elif valve_info[2] == 3:
	
		#Green Boundary valve
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 3)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 10, marker = valve_marker, color = 'green', edgecolors = 'k', zorder = 10)
		
		
	elif valve_info[2] in [4, 5, 6]:
	
		#White Boundary Valve
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 3)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 10, marker = valve_marker, color = 'white', edgecolors = 'k', zorder = 10)
		
	elif valve_info[2] == 7:
	
		#Red Boundary Valve
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 3)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 10, marker = valve_marker, color = 'red', edgecolors = 'k', zorder = 10)

		
	elif valve_info[2] == 8:
	
		#Orange Boundary Valve
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 3)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 10, marker = valve_marker, color = 'orange', edgecolors = 'k', zorder = 10)
		
	elif valve_info[2] == 15:
	
		#Reflux (Non-Return)
		#Similar to washout but much less frequent.
		#Description of formulas can be found below in 'washout'
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 6)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		triangle_marker = MarkerStyle(">")
		triangle_marker._transform.rotate_deg(valve_info[4])
		
		try:
			x_d = 1/sqrt(1+tan(valve_info[4]*pi/180)**2)
			
		except ZeroDivisionError:
			x_d = 0
			
		try:
			y_d = 1/sqrt(1+(1/tan(valve_info[4]*pi/180)**2))
			
		except ZeroDivisionError:
			y_d = 0
		
		
		if valve_info[4] <= 90:
			pass
			
		elif valve_info[4] <=180:
		
			x_d = -1 * x_d
			
		elif valve_info[4] <= 270:
		
			x_d = -1 * x_d
			y_d = -1 * y_d
			
		elif valve_info[4] <= 360:
		
			y_d = -1 * y_d
			
		ax.scatter(valve_info[0] + x_d, valve_info[1] + y_d, s = 1, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
		ax.scatter(valve_info[0] - x_d, valve_info[1] - y_d, s = 50, marker = triangle_marker, color = 'xkcd:hot pink', zorder = 10)
		
		
	elif valve_info[2] == 17:
	
		#Sensistive
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 4)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		ax.scatter(valve_info[0], valve_info[1], s = 3, marker = valve_marker, color = 'gray', zorder = 10)
		
	elif valve_info[2] == 20:
	
		#Washout
		valve_marker = MarkerStyle("s")
		valve_marker._transform.scale(1, 3)
		valve_marker._transform.rotate_deg(valve_info[4])
		
		#The washout symbol is 2 rectangles which need to be parallel and separated by a distance
		#In order for the rectangles to remain parallel they must be moved in the same direction the 
		# main is pointing.
		#These formulas ensure the rectangles are moved correctly with 1.5 ideally moving them 1.5m apart.
		# However, matplotlib markers do not use a physical scale so the distance they are separated depends on the 
		# x and y axis limits (ax.set_xlim and ax.set_ylim). For the limits specified in this code 1.5 works well. 
		try:
			x_d = 1.5/sqrt(1+tan(valve_info[4]*pi/180)**2)
			
		except ZeroDivisionError:
			x_d = 0
			
		try:
			y_d = 1.5/sqrt(1+(1/tan(valve_info[4]*pi/180)**2))
			
		except ZeroDivisionError:
			y_d = 0
		
		#Ensure angle is correct
		if valve_info[4] < 0:
			angle = 180 + valve_info[4]
			
		else:
			angle = valve_info[4]
			
		#Plot rectangles
		if angle <= 90:
			ax.scatter(valve_info[0] + x_d, valve_info[1] + y_d, s = 4, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
			ax.scatter(valve_info[0] - x_d, valve_info[1] - y_d, s = 4, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
			
		else:
			ax.scatter(valve_info[0] - x_d, valve_info[1] + y_d, s = 4, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
			ax.scatter(valve_info[0] + x_d, valve_info[1] - y_d, s = 4, marker = valve_marker, color = 'xkcd:hot pink', zorder = 10)
			
			
			
	#Add a pink circle around the valve if it is closed (and not a washout)
	if valve_info[3] == 15 and not valve_info[2] == 20:
	
		valve_marker = MarkerStyle("o")
		
		if valve_info[2] in [3, 4, 5, 6, 7, 8]:
			ax.scatter(valve_info[0], valve_info[1], s = 120, marker = valve_marker, color = 'none', edgecolors = 'xkcd:hot pink', zorder = 10)
		else:
			ax.scatter(valve_info[0], valve_info[1], s = 80, marker = valve_marker, color = 'none', edgecolors = 'xkcd:hot pink', zorder = 10)

#Plot hydrant info
def plot_hydrant(ax, hydrant_info):

	#Plotting hydrants according to their GIS symbology
	#More information can be found at 'codes\Hydrants\Hydrant Colours.txt'
	if hydrant_info[2] in [2, 3, 7, 8, 9, 10]:
	
		#Default
		marker = MarkerStyle("o")
		
		ax.scatter(hydrant_info[0], hydrant_info[1], s = 20, marker = marker, color = 'xkcd:hot pink', zorder = 10)

	elif hydrant_info[2] == 4:
	
		#Isolated
		marker = MarkerStyle("o")
		
		ax.scatter(hydrant_info[0], hydrant_info[1], s = 20, marker = marker, color = 'purple', zorder = 10)

#Plot PRV info
def plot_prv(ax, prv_info):

	#Plotting PRVs according to their GIS symbology
	#More information can be found at 'codes\PRVs\PRV Symbols.jpg'
	if prv_info[2] == 3:
	
		#Pressure Reducing
		marker = MarkerStyle("s")
		marker._transform.rotate_deg(prv_info[3])
		ax.scatter(prv_info[0], prv_info[1], s = 50, marker = marker, color = 'none', edgecolors = 'xkcd:hot pink', zorder = 10)
		

		marker = MarkerStyle(">")
		marker._transform.rotate_deg(prv_info[3])
		ax.scatter(prv_info[0], prv_info[1], s = 50, marker = marker, color = 'xkcd:hot pink', zorder = 10)
		
	else:
	
		#Default
		#Some symbols are not included as they are rare and if further information is needed GIS will be used 
		marker = MarkerStyle("o")
			
		ax.scatter(prv_info[0], prv_info[1], s = 70, marker = marker, color = 'none', edgecolors = 'xkcd:hot pink', zorder = 10)
			

		marker = MarkerStyle(">")
		marker._transform.rotate_deg(prv_info[3])
		
		ax.scatter(prv_info[0], prv_info[1], s = 50, marker = marker, color = 'xkcd:hot pink', zorder = 10)



#Class to create the matplotlib figure .png and create the .xlsx and .csv outputs
class Map():

	#Start creating of .png using non run specific data 
	def __init__(self, directory, checker, mains):
	
		#checker needed for data sanitation
		self.checker = checker
		
		#directory is DMA dependent so much be passed in
		self.directory = directory
	
		#dma_outline with a buffer of 10m
		dma_outline = checker.get_polygon().buffer(10)
	
	
	
	
	
		############################################################################################
		#Data collection from .shp files
		############################################################################################
	
		#Gathering and sanitising address data
		address_data = address_info(dma_outline)		
		address_data = data_within_outline(address_data, checker)
		
		#Making address data accessible to get function
		self.address_data = address_data
		
		
		#Creating array with address xy coordinates only
		address_xy = []
		for i in range(len(address_data)):
			address_xy.append([address_data[i][0], address_data[i][1]])
		
		
		#Gathering and sanitising valve data
		valve_data = valve_info(dma_outline)
		valve_data = data_within_outline(valve_data, checker)
	
		#Creating list of boundary valves needed in grouping.py
		boundary_valves = []
		for i in range(len(valve_data)):
		
			if valve_data[i][2] in [3, 4, 5, 6, 7, 8]:
			
				boundary_valves.append(valve_data[i])
	
		self.boundary_valves = boundary_valves
	
		#Gathering and sanitising hydrant data
		hydrant_data = hydrant_info(dma_outline)
		hydrant_data = data_within_outline(hydrant_data, checker)
	
		#Gathering and sanitising PRV data
		prv_data = prv_info(dma_outline)
		prv_data = data_within_outline(prv_data, checker)
	

	
		############################################################################################
		#Matplotlib Plotting
		############################################################################################
	
		#Creating figure
		#figsize has the same aspect ratio as 16:9
		fig = plt.figure(1, figsize = (6.4, 3.6))
		ax = fig.add_subplot()
		
		
		#Water mains data created in grouping.py
		m_geo = mains[0]
		p_geo = mains[1]
		o_geo = mains[2]

		#Plotting water mains
		p_geo.plot(ax = ax, color = 'xkcd:light blue')
		m_geo.plot(ax = ax, color = 'blue')
		o_geo.plot(ax = ax, color = 'black')
		
		#Plotting address data
		for i in range(len(address_xy)):
			ax.scatter(address_xy[i][0], address_xy[i][1], s = 2, color = 'black')
		
		#Plotting valve data
		for i in range(len(valve_data)):
			plot_valve(ax, valve_data[i])
			
		#Plotting hydrant data	
		for i in range(len(hydrant_data)):
			plot_hydrant(ax, hydrant_data[i])
			
		#Plotting PRV data
		for i in range(len(prv_data)):
			plot_prv(ax, prv_data[i])
		
		
		############################################################################################
		#Background Map Plotting
		############################################################################################
		
		#Polygon == DMA outline
		poly = checker.get_polygon()
		self.poly = poly
		
		#Extent of DMA
		extent = poly.bounds
		
		#Getting titles of background maps which contain the DMA
		map_strings = MapStrings()
		map_string_list = map_strings.get_map_strings(extent)
		
		#Directory where background maps are stored
		folder = r'.shp\Maps\Maps'
		
		#Loop to display every map in figure - memory intensive
		for i in range(len(map_string_list)):
		
			#letters refer to sub-folder in '.shp\Maps\Maps'
			letters = map_string_list[i][0] + map_string_list[i][1]
			
			#Name of map file
			file = '{0}\\{1}\\{2}.tif'.format(folder, letters, map_string_list[i])
			
			#Use georaster to keep colour scale as well as extent
			my_image = georaster.SingleBandRaster(file)
			
			#Read and plot background map file
			img=mpimg.imread(file) #Most memory intensive line 
			plt.imshow(img, extent=my_image.extent, alpha = 0.6)
		
		
		
		#Make fig and ax accessible for more information to be added
		self.fig = fig
		self.ax = ax
		
		
		#plt.show()
		
		
	def add_features(self, filename, feature_data):
	
	
		directory = self.directory
	
		fig = self.fig
		ax = self.ax
	
		#Data from other modules to be plotted
		algorithm_data = feature_data[0]
		final_POIs  = feature_data[1]
		poi_index = feature_data[2]
		nearest_address_index = feature_data[3]
		data_list = feature_data[4]
		poi_id_list = feature_data[5]
	
		address_data = self.address_data
		
		checker = self.checker
		
		############################################################################################
		#Data Plotting
		############################################################################################
		
		#List to hold run specific data to be deleted once figure is saved
		plot_data_to_remove = []
		
		#Plotting non POIs in green
		for i in range(len(algorithm_data)):
			if algorithm_data[i][3] <= 1:
				plot_data_to_remove.append(ax.scatter(algorithm_data[i][0], algorithm_data[i][1], s = 25, color = 'xkcd:bright green', edgecolors='k', zorder = 10))
				
		#Plotting group POIs in dark red, sub POIs in orange and normal POIs in red
		for i in range(len(final_POIs)):
		
			if len(final_POIs[i][3]) > 0:
			
				plot_data_to_remove.append(ax.scatter(final_POIs[i][0], final_POIs[i][1], color = 'xkcd:scarlet', s = 40, edgecolors='k', zorder = 12))
				
				for j in range(len(final_POIs[i][3])):
				
					plot_data_to_remove.append(ax.scatter(poi_index[int(final_POIs[i][3][j])][1][0], poi_index[int(final_POIs[i][3][j])][1][1], s = 40, color = 'xkcd:bright orange', edgecolors='k', zorder = 11))

			else:
			
				plot_data_to_remove.append(ax.scatter(final_POIs[i][0], final_POIs[i][1], s = 40, color = 'xkcd:bright red', edgecolors='k', zorder = 12))
				

		#Plotting polygon to show DMA outline
		poly = self.poly
		try:
			ax.plot(*poly.exterior.xy, linewidth = 0.5, color = 'xkcd:bright green')
		
		except AttributeError:
			for i in range(len(poly)):
				ax.plot(*poly[i].exterior.xy, linewidth = 0.5, color = 'xkcd:bright green')
	
		
		#This loop separates the POI and sub POI data from 'data_list'
		poi_data = []
		sub_poi_data = []
		for i in range(len(poi_id_list)):
		
			if not str(poi_id_list[i]) == '':
			
				#Combining POI data with POI ID
				poi_data.append([float(poi_id_list[i]), float(data_list[i][0]), float(data_list[i][1]), data_list[i][2], float(data_list[i][3]), data_list[i][4], data_list[i][5], data_list[i][6], float(data_list[i][7])])
		
				j = i + 1
				
				add_pois = True
				sub_poi_array = []
				
				
				if j >= len(data_list):
				
					add_pois = False
				
				
				#Creating list to hold the sub POI data and link all sub POIs under one POI ID
				while add_pois:
				
					#Adding sub POIs to the list until the next POI in 'data_list' is reached
					if data_list[j][2] == 'sub-POI':
				
						sub_poi_array.append([float(data_list[j][0]), float(data_list[j][1]), float(data_list[j][3]), data_list[j][4], data_list[j][5]])
						
						j += 1
						
						if j >= len(data_list):
						
							add_pois = False
						
					else:
					
						add_pois = False
	
				sub_poi_data.append([poi_id_list[i], sub_poi_array])
	

		#Loop to save a figure showing every POI in detail
		for i in range(len(poi_data)):
			
			#Plot purple star indicating nearest address
			ax.scatter(address_data[nearest_address_index[i]][0], address_data[nearest_address_index[i]][1], marker = '*', s = 40, color = 'xkcd:bright purple')
			
			#Set xy scale - making sure to keep aspect ratio 16:9
			ax.set_xlim(poi_data[i][1] - 176, poi_data[i][1] + 176)
			ax.set_ylim(poi_data[i][2] - 99, poi_data[i][2] + 99)
		
			#No whitespace around figure
			plt.axis('off')
			fig.tight_layout(pad=0, w_pad=0, h_pad=0)
			
			#Specifiying title and directory
			plot_folder = r'lib\png'
			plt_title = '{0}\\{1}.png'.format(plot_folder, str(int(poi_data[i][0])))
			
			#Saving figure
			plt.savefig(plt_title)
			
		#Remove all run specific data so that figure can be reused
		for d in plot_data_to_remove:
			d.remove()
		
		
		############################################################################################
		#.csv Creation for GIS
		############################################################################################
		

		
		#Adding entry in 'poi_id_list' for non POIs so that 'poi_id_list' will have the same length as 'data_list' 
		for i in range(len(poi_id_list), len(data_list)):
		
			poi_id_list.append('')

		#Creating dict to hold data for GIS output
		#try:	
		gis_data = {'POI_ID' : poi_id_list, 'X' : data_list[:,0], 'Y' : data_list[:,1], 'Point Type' : data_list[:,2], 'Confidence' : data_list[:,3]}

		#If no POIs are found 
		#except IndexError:
		#	gis_data = {'POI_ID' : ['No POIs found'], 'X' : [''], 'Y' : [''], 'Point Type' : [''], 'Confidence' : ['']}

		#Creating .csv for GIS upload
		gis_frame = pd.DataFrame(gis_data)
		gis_frame.to_csv('{0}\\GIS\\{1} - GIS.csv'.format(directory, filename), index = False)
		
		
		############################################################################################
		#.xlsx Creation
		############################################################################################
		
		
		#Creating workbook and naming sheets
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = 'POI Overview'
		ws2 = wb.create_sheet(title="sub-POI Information")
		
		
		#Sub POI worksheet column names
		sub_POI_titles = ['POI_ID', 'X', 'Y', 'Confidence', 'Logger Profile', 'Material']
		
		for i in range(len(sub_POI_titles)):
		
			ws2.cell(column=i + 1, row=1, value=sub_POI_titles[i])
			
			
		row_counter = 2
		
		#Add sub POI information to the spreadsheet
		for i in range(len(sub_poi_data)):
		
			if len(sub_poi_data[i][1]) <= 1:
				continue
		
			#Cell manipulation for better appearance
			ws2.merge_cells(start_row = row_counter, start_column = 1, end_row = row_counter + len(sub_poi_data[i][1]) - 1, end_column = 1)
			ws2.cell(column = 1, row = row_counter).alignment = Alignment(horizontal="center", vertical="center")
			
			#Add POI ID
			ws2.cell(column = 1, row = row_counter, value = poi_data[i][0])
		
			#Add sub POI data
			for j in range(len(sub_poi_data[i][1])):
				
				for k in range(len(sub_poi_data[i][1][j])):
				
					ws2.cell(column = 2 + k, row = row_counter + j, value = sub_poi_data[i][1][j][k])
		
			row_counter += len(sub_poi_data[i][1])
		
		
		#Resizing column widths for better appearance
		column_width = [8, 10, 10, 12, 14, 10] 
		i = 0
		for column_cells in ws2.columns:
		
			length = column_width[i]

			ws2.column_dimensions[column_cells[0].column_letter].width = length
			
			i += 1
		
		


		if len(poi_data) < 1:
		
			#Cell manipulation for better appearance
			ws.merge_cells(start_row = 2, start_column = 2, end_row = 2, end_column = 3)
			ws.cell(column = 2, row = 2).alignment = Alignment(horizontal="center", vertical="center")
			
			#If no POIs have been found then display message
			ws.cell(column = 2, row = 2, value = 'No POIs Found')
			
			
		else:
		
			#'populate_spreadsheet' packages the POI information so that all POI data can be added with respect
			# to an anchor cell. This cell will be updated after every POI is displayed
			anchor_cell = [1,1]
			for i in range(len(poi_data)):
				anchor_cell = populate_spreadsheet(wb, poi_data[i], anchor_cell)
		
		
		#Make sure filename ends with .xlsx
		if not filename.endswith('.xlsx'):
			filename = '{0}.xlsx'.format(filename)


		#Save workbook to output folder
		output_folder = "{0}\\{1}".format(directory, filename)
		wb.save(output_folder)
		

		
		
	#Return address info
	def get_address_data(self):
	
		return self.address_data
		
	#Return boundary valve info
	def get_bv_data(self):
	
		return self.boundary_valves
		
	#Close figure so new plot can be made
	def close_fig(self):
	
		plt.close()

		
